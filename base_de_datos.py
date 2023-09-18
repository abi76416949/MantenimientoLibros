import openpyxl
from openpyxl import Workbook

class BaseDeDatos:
    def __init__(self, archivo):
        self.archivo = archivo

    #---------------------ABRIR LIBROS--------------------------------
    def _abrir_archivo(self):
        try:
            libro = openpyxl.load_workbook(self.archivo)
            return libro
        except FileNotFoundError:
            return None
        
    #---------------------------OBTENER_LIBROS------------------------------------
    def obtener_libros(self):

        libro, hoja = self._abrir_archivo()
        
        if hoja is None:
            return []

        datos = []
        for row in hoja.iter_row(values_only=True):
            if row[0] != 'codigo_libro':
                datos.append(
                    {'codigo_libro': row[0], 'titulo': row[1], 'aho': row[2], 'tomo': row[3]})
        libro.close()
        return datos
    #------------------------GUARDAR_lIBROS--------------------------------------
    def guardar_libros(self, libros):
        libro = Workbook()
        hoja = libro.active
        hoja.append(['codigo_libro', 'titulo', 'aho', 'tomo'])
        for libro_data in libros:
            hoja.append([libro_data['codigo_libro'], libro_data['titulo'], libro_data['aho'], libro_data['tomo']])        
            libro.save(self.archivo)


    #----------------------OBTENER-CATEGORIAS------------------------------------
    def obtener_categorias(self):

        libro, hoja = self._abrir_archivo()
        if hoja is None:
            return []
        
        datos = []
        for row in hoja.iter_row(values_only=True):
            if row[0] != 'cod_categoria':
                datos.append({'cod_categoria': row[0], 'categoria': row[1]})
        libro.close()
        return datos

    # -----------------GUARDAR CATEGORIAS--------------------------------------
    def guardar_categorias(self, categorias):
        libro, hoja = self._abrir_archivo()
        if hoja is None:
            libro = Workbook()
            hoja = libro.active
            hoja.append(['cod_categoria', 'categoria'])

        for categoria in categorias:
            #Genera el codigo aleatorio de categoria antes de guardar el archivo
            codigo_categoria = categoria.generar_codigo_categoria()
            hoja.append([codigo_categoria, categoria.categoria])

        libro.save(self.archivo)

    # --------------------------OBTENER PERSONAS----------------------------------------------------
    def obtener_personas(self):

        libro, hoja = self._abrir_archivo()

        if hoja is None:
            return []
        datos = []
        for row in hoja.iter_row(values_only=True):
            if row[0] != 'cod_persona':
                datos.append({'cod_persona': row[0], 'nombre': row[1], 'apellidoPaterno': row[2],
                            'apellidoMaterno': row[3], 'fecha_nacimieto': row[4]})
        libro.close()
        return datos

    # -----------------GUARDAR PERSONAS--------------------------------------
    def guardar_personas(self, personas):
        libro = Workbook()
        hoja = libro.active
        hoja.append(['cod_persona', 'nombre', 'apellidoPaterno',
                    'apellidoMaterno', 'fecha_nacimieto'])
        for persona in personas:
            hoja.append([persona.cod_persona, persona.nombre, persona.apellidoPaterno,
                        persona.apellidoMaterno, persona.fecha_nacimieto])
        libro.save(self.archivo)
    #---------- AAGREGAR CATEGORIAS--------------------------------------
    def agregar_categorias(self):
        self.cod_categoria = input(print("agrega el codi de la categoria"))
        self.categoria = input(print("agrega el nombre de la categoria"))
        


    #-----------ELIMINAR LIBROS-------------------------------------------
    def eliminar_libro(self, codigo_libro):
        libro, hoja= self._abrir_archivo()
        fila_a_eliminar = None
        for row in hoja.iter_rows(values_only=True):
            if row[0] == codigo_libro:
                fila_a_eliminar = row
                break

        if fila_a_eliminar:
            hoja.delete_rows(hoja.index(fila_a_eliminar[0]))

                # Guardar los cambios en el archivo
            libro.save(self.archivo)
            libro.close()


    #----------------Editar libro-------------------------------------------
    def editar_libro(self, codigo_libro, nuevo_titulo, nuevo_ano, nuevo_tomo):
        libro, hoja = self._abrir_archivo()
        fila_a_editar = None

        for row in hoja.iter_rows(values_only=True):
            if row[0] == codigo_libro:
                fila_a_editar = row
                break

        if fila_a_editar:
            # Actualiza los valores del libro
            fila_a_editar[1] = nuevo_titulo
            fila_a_editar[2] = nuevo_ano
            fila_a_editar[3] = nuevo_tomo

            # Guarda los cambios en el archivo
            libro.save(self.archivo)
            libro.close()
            return True  # Libro encontrado y editado
        else:
            libro.close()
            return False  # Libro no encontrado
        
    #-----------------BUSCAR LIBROS-------------------------------------------
    def buscar_libro(self, codigo_libro):
        libro, hoja = self._abrir_archivo()
        fila_a_buscar = None
        for row in hoja.iter_rows(values_only=True):
            if row[0] == codigo_libro:
                fila_a_buscar = row
                break

        if fila_a_buscar:
            libro.close()
            # Retorna la información del libro encontrado
            return {
                'codigo_libro': fila_a_buscar[0],
                'titulo': fila_a_buscar[1],
                'aho': fila_a_buscar[2],
                'tomo': fila_a_buscar[3]
            }
        else:
            libro.close()
            return None
        
    ######################AUTORES GESTION #######################################


#----------------AGREGAR AUTORES-----------------------------------------------------------
    def agregar_autores(self, autores):
        libro, hoja = self._abrir_archivo()
        if hoja is None:
            libro = Workbook()
            hoja = libro.active
            hoja.append(['codigo_autor', 'nombre', 'apellidoMaterno', 'apellidoPaterno', 'pais', 'editorial'])

        for autor_data in autores:
            hoja.append([autor_data['codigo_autor'], autor_data['nombre'], autor_data['apellidoMaterno'], autor_data['apellidoPaterno'], autor_data['pais'], autor_data['editorial']])

        libro.save(self.archivo)
        libro.close()

#------------ASIGNAR UN AUTOR A UN LIBRO-------------------------------------------------
    def asignar_autor_a_un_libro(self,codigo_libro, codigo_autor):
        libro, hoja = self._abrir_archivo()
        if hoja is None:
            libro = Workbook()
            hoja = libro.active
            hoja.append(['codigo_libro', 'codigo_autor'])

        hoja.append([codigo_libro, codigo_autor])
        libro.save(self.archivo)
        libro.close()

#---------------OBTENER AUTORES-------------------
    def obtener_autores(self):
        libro, hoja = self._abrir_archivo()
        if hoja is None:
            return []

        datos = []
        for row in hoja.iter_row(values_only=True):
            if row[0] != 'codigo_autor':
                datos.append({'codigo_autor': row[0], 'nombre': row[1], 'apellidoMaterno': row[2], 'apellidoPaterno': row[3], 'pais': row[4], 'editorial': row[5]})
        libro.close()
        return datos
    
#-----------------Eliminar AUTORES--------------------------------------
    def eliminar_autor(self, codigo_autor):
        libro, hoja= self._abrir_archivo()
        fila_a_eliminar = None
        for row in hoja.iter_rows(values_only=True):
            if row[0] == codigo_autor:
                fila_a_eliminar = row
                break

        if fila_a_eliminar:
            hoja.delete_rows(hoja.index(fila_a_eliminar[0]))

                # Guardar los cambios en el archivo
            libro.save(self.archivo)
            libro.close()

#----------Editar AUTORES---------------------------------------------------
    def editar_autores(self, codigo_autor, nuevo_nombre, nuevo_apellidoMaterno,nuevo_apellidoPaterno, nuevo_pais, nuevo_editorial):
        libro, hoja = self._abrir_archivo()
        fila_a_editar = None

        for row in hoja.iter_rows(values_only=True):
            if row[0] == codigo_autor:
                fila_a_editar = row
                break

        if fila_a_editar:
            # Actualiza los valores del libro
            fila_a_editar[1] = nuevo_nombre
            fila_a_editar[2] = nuevo_apellidoMaterno
            fila_a_editar[3] = nuevo_apellidoPaterno
            fila_a_editar[4] = nuevo_pais
            fila_a_editar[5] = nuevo_editorial

            # Guarda los cambios en el archivo
            libro.save(self.archivo)
            libro.close()
            return True
    
    def cerrar(self):
        pass
    